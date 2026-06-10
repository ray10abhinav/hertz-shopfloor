from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
from datetime import date
import asyncpg, os, io, re
from dotenv import load_dotenv

load_dotenv()
app = FastAPI(title="Hertz Shopfloor API")

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

DB_URL = os.getenv("DATABASE_URL", "postgresql://postgres:password@localhost:5432/hertz")
pool: asyncpg.Pool = None

@app.on_event("startup")
async def startup():
    global pool
    pool = await asyncpg.create_pool(DB_URL)

@app.on_event("shutdown")
async def shutdown():
    await pool.close()

# ── Models ────────────────────────────────────────────────────
class PlanRow(BaseModel):
    date: date
    line: int
    supervisor: str
    job_no: str
    product: str
    ml: float = 100
    target_qty: int
    job_type: str = "Filling & Packing"

class ActualUpsert(BaseModel):
    date: date
    line: int
    job_no: str
    actual_qty: Optional[int] = None   # None = not produced
    not_produced: bool = False
    period_from: Optional[str] = None
    period_to: Optional[str] = None
    batch_no: Optional[str] = None
    reason: Optional[str] = None
    remark: Optional[str] = None

class AddonCreate(BaseModel):
    date: date
    line: int
    product: str
    job_no: Optional[str] = None
    ml: float = 100
    actual_qty: int = 0
    period_from: Optional[str] = None
    period_to: Optional[str] = None
    batch_no: Optional[str] = None
    reason: Optional[str] = None
    remark: Optional[str] = None

class SubmitLine(BaseModel):
    date: date
    line: int

# ── Plan endpoints ─────────────────────────────────────────────
@app.get("/plan")
async def get_plan(date: date, line: Optional[int] = None):
    if line:
        rows = await pool.fetch("SELECT * FROM plan WHERE date=$1 AND line=$2 ORDER BY id", date, line)
    else:
        rows = await pool.fetch("SELECT * FROM plan WHERE date=$1 ORDER BY line, id", date)
    return [dict(r) for r in rows]

@app.get("/plan/dates")
async def get_plan_dates():
    rows = await pool.fetch("SELECT DISTINCT date FROM plan ORDER BY date")
    return [str(r["date"]) for r in rows]

@app.post("/plan/bulk")
async def upload_plan(rows: List[PlanRow]):
    inserted = 0
    errors = []
    async with pool.acquire() as conn:
        async with conn.transaction():
            for r in rows:
                try:
                    await conn.execute(
                        """INSERT INTO plan (date, line, supervisor, job_no, product, ml, target_qty, job_type)
                           VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
                           ON CONFLICT (date, line, job_no)
                           DO UPDATE SET supervisor=EXCLUDED.supervisor, product=EXCLUDED.product,
                                         ml=EXCLUDED.ml, target_qty=EXCLUDED.target_qty,
                                         job_type=EXCLUDED.job_type""",
                        r.date, r.line, r.supervisor, r.job_no,
                        r.product, r.ml, r.target_qty, r.job_type
                    )
                    inserted += 1
                except Exception as e:
                    errors.append(f"{r.job_no}: {e}")
    return {"inserted": inserted, "errors": errors[:10]}

@app.post("/plan/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...)):
    import openpyxl
    from datetime import datetime as dt
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {str(h).strip().lower(): i for i, h in enumerate(raw_headers) if h is not None}
    def gc(keys):
        for k in keys:
            if k.lower() in col: return col[k.lower()]
        return None
    rows, skipped = [], []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row): continue
        try:
            raw_date = row[gc(["date"])]
            if raw_date is None: skipped.append(f"Row {i}: missing date"); continue
            d = raw_date.date() if hasattr(raw_date, "date") else dt.strptime(str(raw_date)[:10], "%Y-%m-%d").date()
            raw_line = row[gc(["line no.", "line no", "line"])]
            try: line = int(raw_line)
            except (TypeError, ValueError): skipped.append(f"Row {i}: non-numeric line '{raw_line}' skipped"); continue
            supervisor = str(row[gc(["supervisor"])] or "").strip()
            job_no = str(row[gc(["job no.", "job no"])] or "").strip()
            product = str(row[gc(["product name", "product"])] or "").strip()
            ml = float(row[gc(["ml"])] or 100)
            tq = row[gc(["target qty.", "target qty", "target"])]
            if tq is None: target_qty = 0
            elif not isinstance(tq, (int, float)):
                m = re.match(r"[\d,]+", str(tq).replace(",", ""))
                target_qty = int(m.group()) if m else 0
            else: target_qty = int(tq)
            job_type = str(row[gc(["job"])] or "Filling & Packing").strip()
            if not job_no or not product or line == 0: skipped.append(f"Row {i}: missing data"); continue
            rows.append(PlanRow(date=d, line=line, supervisor=supervisor, job_no=job_no, product=product, ml=ml, target_qty=target_qty, job_type=job_type))
        except Exception as e: skipped.append(f"Row {i}: {e}"); continue
    if not rows: raise HTTPException(400, f"No valid rows. Issues: {skipped[:5]}")
    result = await upload_plan(rows)
    result["skipped"] = len(skipped)
    result["skip_reasons"] = skipped[:10]
    return result

# ── Actuals endpoints ──────────────────────────────────────────
@app.get("/actuals")
async def get_actuals(date: date, line: Optional[int] = None):
    if line:
        rows = await pool.fetch("SELECT * FROM actuals WHERE date=$1 AND line=$2", date, line)
        addons = await pool.fetch("SELECT * FROM actuals_addon WHERE date=$1 AND line=$2", date, line)
    else:
        rows = await pool.fetch("SELECT * FROM actuals WHERE date=$1", date)
        addons = await pool.fetch("SELECT * FROM actuals_addon WHERE date=$1", date)
    return {"actuals": [dict(r) for r in rows], "addons": [dict(r) for r in addons]}

@app.put("/actuals")
async def upsert_actual(body: ActualUpsert):
    await pool.execute(
        """INSERT INTO actuals (date, line, job_no, actual_qty, not_produced, period_from, period_to, batch_no, reason, remark, updated_at)
           VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,NOW())
           ON CONFLICT (date, line, job_no)
           DO UPDATE SET actual_qty=EXCLUDED.actual_qty, not_produced=EXCLUDED.not_produced,
                         period_from=EXCLUDED.period_from, period_to=EXCLUDED.period_to,
                         batch_no=EXCLUDED.batch_no, reason=EXCLUDED.reason,
                         remark=EXCLUDED.remark, updated_at=NOW()""",
        body.date, body.line, body.job_no,
        None if body.not_produced else body.actual_qty,
        body.not_produced,
        body.period_from, body.period_to,
        body.batch_no, body.reason, body.remark
    )
    return {"ok": True}

@app.post("/actuals/addon")
async def add_addon(body: AddonCreate):
    row = await pool.fetchrow(
        """INSERT INTO actuals_addon (date, line, product, job_no, ml, actual_qty, period_from, period_to, batch_no, reason, remark)
           VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11) RETURNING id""",
        body.date, body.line, body.product, body.job_no, body.ml, body.actual_qty,
        body.period_from, body.period_to, body.batch_no, body.reason, body.remark
    )
    return {"id": row["id"]}

@app.delete("/actuals/addon/{addon_id}")
async def delete_addon(addon_id: int):
    await pool.execute("DELETE FROM actuals_addon WHERE id=$1", addon_id)
    return {"ok": True}

@app.post("/actuals/submit")
async def submit_line(body: SubmitLine):
    await pool.execute("UPDATE actuals SET submitted=TRUE WHERE date=$1 AND line=$2", body.date, body.line)
    return {"ok": True}

# ── Report: plan JOIN actuals (the "Filling Data" view) ────────
@app.get("/report")
async def get_report(date: date, line: Optional[int] = None):
    """Returns plan vs actual joined rows with delta calculated."""
    if line:
        rows = await pool.fetch("SELECT * FROM report WHERE date=$1 AND line=$2 ORDER BY line, job_no", date, line)
    else:
        rows = await pool.fetch("SELECT * FROM report WHERE date=$1 ORDER BY line, job_no", date)
    return [dict(r) for r in rows]

# ── Dashboard summary ──────────────────────────────────────────
@app.get("/dashboard")
async def dashboard(date: date):
    plan_rows = await pool.fetch(
        "SELECT line, supervisor, SUM(target_qty) AS planned FROM plan WHERE date=$1 GROUP BY line, supervisor ORDER BY line", date)
    actual_rows = await pool.fetch(
        "SELECT line, SUM(actual_qty) AS actual, BOOL_OR(submitted) AS submitted FROM actuals WHERE date=$1 GROUP BY line", date)
    addon_rows = await pool.fetch(
        "SELECT line, COUNT(*) AS addon_count, SUM(actual_qty) AS addon_qty FROM actuals_addon WHERE date=$1 GROUP BY line", date)
    actual_map = {r["line"]: r for r in actual_rows}
    addon_map  = {r["line"]: r for r in addon_rows}
    result = []
    for p in plan_rows:
        line = p["line"]
        a = actual_map.get(line)
        ao = addon_map.get(line)
        planned = int(p["planned"])
        actual  = int(a["actual"]) if a and a["actual"] else 0
        result.append({
            "line": line, "supervisor": p["supervisor"],
            "planned": planned, "actual": actual,
            "delta": actual - planned,
            "delta_pct": round((actual - planned) / planned * 100, 1) if planned else None,
            "submitted": bool(a["submitted"]) if a else False,
            "addon_count": int(ao["addon_count"]) if ao else 0,
        })
    return result
